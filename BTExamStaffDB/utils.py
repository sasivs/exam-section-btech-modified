from openpyxl import Workbook
from openpyxl import styles


class FacultyInfoTemplateBookGenerator:
    def __init__(self):
        self.fields = ['FacultyId', 'Name', 'Phone', 'Email', 'Dept']

    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_template_sheet(workbook=workbook)
        return workbook
    
    def generate_template_sheet(self, workbook):
        file = 'FacultyInfo-Template'
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        row_num = 1
        for col_num, column_title in enumerate(self.fields,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title