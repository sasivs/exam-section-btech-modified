from openpyxl import Workbook
from openpyxl import styles

class SampleMarksUploadExcelSheetGenerator:
    def __init__(self, students, regEvent, subject):
        self.students = students
        self.regEvent = regEvent
        self.subject = subject
    
    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_marks_upload_sheet(workbook=workbook)
        return workbook

    def generate_marks_upload_sheet(self, workbook):
        file = self.regEvent.__str__()
        file = file.replace(':', '-')
        file += '__'+str(self.subject.SubCode)
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        marks_distribution = self.subject.MarkDistribution
        marks_distribution_names = marks_distribution.split(',')
        marks_distribution_names = [name.split('+') for name in marks_distribution_names]
        marks_distribution_names = ['RegNo', 'Name'] + marks_distribution_names
        row_num = 1
        for col_num, column_title in enumerate(marks_distribution_names,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title
        
        for student in self.students:
            row_num+=1
            row_data = [
                student.RegNo,
                student.Name
            ]
            empty_cols = ['' for col in range(len(marks_distribution_names)-2)]
            row_data = row_data + empty_cols
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
