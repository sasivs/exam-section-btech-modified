from openpyxl import Workbook, styles
from openpyxl.worksheet.datavalidation import DataValidation

from BTsuperintendent.models import BTMarksDistribution

class CoursesTemplateExcelFile:

    def __init__(self, regulation, byear):
        self.byear = byear
        self.regulation = regulation
        self.mark_distribution_objs = BTMarksDistribution.objects.filter(Regulation=self.regulation)
    
    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_course_sheet(workbook=workbook)
        return workbook

    def generate_course_sheet(self, workbook):
        file = str(self.regulation) + '_' + str(self.byear)
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        headers = ['CourseCode', 'CourseName', 'BYear', 'BSem', 'Department', 'OfferedBy', 'Regulation',\
            'Creditable', 'Credits', 'Type', 'Category', 'lectures', 'tutorials', 'practicals', 'DistributionRatio', 'MarkDistribution']
        MARK_DISTRIBUTION_CHOICES = [str(dis.Distribution)+'('+str(dis.PromoteThreshold)+str(')') for dis in self.mark_distribution_objs]
        mark_distr_choices_str = ', '.join(MARK_DISTRIBUTION_CHOICES)
        mark_distr_choices_str = str('\"')+mark_distr_choices_str+str('\"')
        data_val = DataValidation(type="list", formula1=mark_distr_choices_str, allow_blank=False)
        worksheet.add_data_validation(data_val)
        row_num = 1
        for col_num, column_title in enumerate(headers,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title
        col_num = 16
        for row_num in range(2,53):
            cell = worksheet.cell(row=row_num, column=col_num)
            data_val.add(cell)
        
