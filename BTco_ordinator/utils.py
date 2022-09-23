from openpyxl import Workbook
from openpyxl import styles
from BTco_ordinator.models import BTSubjects

class RollListBookGenerator:
    def __init__(self, rollList, regEventId):
        self.rollList = rollList
        self.regEvent = regEventId
    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_roll_sheet(workbook=workbook)
        return workbook
    
    def generate_roll_sheet(self, workbook):
        file = str(self.regEvent)
        file = file.replace(':', '-')
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        headers = ['id', 'RegNo', 'AYear', 'BYear', 'ASem', 'Bsem', 'Dept', 'Regulation', 'Cycle', 'Section']
        row_num = 1
        for col_num, column_title in enumerate(headers,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title
        
        for roll in self.rollList:
            row_num+=1
            row_data = [
                roll.id,
                roll.student.RegNo,
                roll.RegEventId.AYear,
                roll.RegEventId.BYear,
                roll.RegEventId.ASem,
                roll.RegEventId.BSem,
                roll.RegEventId.Dept,
                roll.RegEventId.Regulation,
                roll.Cycle,
                roll.Section
            ]
            for col_num, cell_value in enumerate(row_data,1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

class NotPromotedBookGenerator:
    def __init__(self, not_promoted_list, regulation, regEventId):
        self.not_promoted = not_promoted_list
        self.regulation = regulation
        self.regEvent = regEventId

    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_not_promoted_sheet(workbook=workbook)
        return workbook
    
    def generate_not_promoted_sheet(self, workbook):
        file = str(self.regEvent)
        file = file.replace(':', '-')
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        headers = ['student_id', 'RegNo', 'RollNo', 'Name', 'Academic Year', 'BTech Year', 'Regulation', 'PoA']
        row_num = 1
        for col_num, column_title in enumerate(headers,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold = True)
            cell.value = column_title
        
        for row in self.not_promoted:
            row_num+=1
            row_data = [
                row['student'].id,
                row['student'].RegNo,
                row['student'].RollNo,
                row['student'].Name,
                row['AYear'],
                row['BYear'],
                row['Regulation'],
                row['PoA']
            ]
            for col_num, cell_value in enumerate(row_data,1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

class ModelTemplateBookGenerator:
    def __init__(self, klass):
        if not hasattr(klass, "_default_manager"):
            klass__name = (
                klass.__name__ if isinstance(klass, type) else klass.__class__.__name__
            )
            raise ValueError(
                "Argument must be a model, not '%s'." % klass__name
            )
        self.fields = [field.name for field in klass._meta.fields]
        self.fields.pop(0)
        self.klass = klass.__name__
        

    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_template_sheet(workbook=workbook)
        return workbook
    
    def generate_template_sheet(self, workbook):
        file = str(self.klass)
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        row_num = 1
        for col_num, column_title in enumerate(self.fields,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title

class SubjectsTemplateBookGenerator:
    def __init__(self):
        self.fields = ['SubCode', 'SubName', 'BYear', 'BSem', 'Dept','OfferedYear', 'Regulation',\
                        'Creditable', 'Credits','Type','Category', 'OfferedBy']

    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_template_sheet(workbook=workbook)
        return workbook
    
    def generate_template_sheet(self, workbook):
        file = 'Subjects'
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        row_num = 1
        for col_num, column_title in enumerate(self.fields,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title

class RegNoTemplateBookGenerator:
    def __init__(self):
        self.fields = ['RegNo']

    def generate_workbook(self):
        workbook = Workbook()
        workbook.remove(workbook.active)

        self.generate_template_sheet(workbook=workbook)
        return workbook
    
    def generate_template_sheet(self, workbook):
        file = 'RegNo-Model'
        worksheet = workbook.create_sheet(title = "{filename}".format(filename=file))
        row_num = 1
        for col_num, column_title in enumerate(self.fields,1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = styles.Font(bold=True)
            cell.value = column_title